from __future__ import annotations



import os

import re

from dataclasses import dataclass

from pathlib import Path

from typing import Dict, Iterable, Optional, Pattern



import pandas as pd

from openpyxl import load_workbook

from openpyxl.utils import get_column_letter

from openpyxl.worksheet.table import Table, TableStyleInfo



TEMPLATE_SHEET = "Template"

RULES_SHEET = "ValidationRules"

CHECK_MARK = "OK"

CROSS_MARK = "KO"





@dataclass

class ValidationRule:

    field: str

    checked: bool

    required: bool

    min_length: Optional[int]

    max_length: Optional[int]

    allowed_values: Optional[set[str]]

    allowed_source: Optional[str]

    pattern: Optional[Pattern[str]]

    custom_rule: Optional[str]





def normalize(value: str) -> str:

    return str(value).strip().lower().replace(" ", "").replace("\n", "")





def parse_bool(value: object) -> bool:

    if isinstance(value, bool):

        return value

    if value is None or (isinstance(value, float) and pd.isna(value)):

        return False

    text = str(value).strip().lower()

    return text in {"true", "1", "yes", "y", "oui", "x"}





def parse_int(value: object) -> Optional[int]:

    if value is None or (isinstance(value, float) and pd.isna(value)):

        return None

    try:

        return int(value)

    except (ValueError, TypeError):

        return None





def load_template(input_file: str) -> pd.DataFrame:

    return pd.read_excel(input_file, sheet_name=TEMPLATE_SHEET)





def load_validation_rules(input_file: str, override: Optional[pd.DataFrame] = None) -> tuple[dict[str, ValidationRule], dict[str, pd.DataFrame]]:
    rules_df = override if override is not None else pd.read_excel(input_file, sheet_name=RULES_SHEET)


    reference_cache: dict[str, pd.DataFrame] = {}

    rules: dict[str, ValidationRule] = {}



    for _, row in rules_df.iterrows():

        field = str(row["Field"]).strip()

        if not field:

            continue



        allowed_source = row.get("AllowedValues") if isinstance(row.get("AllowedValues"), str) else None

        allowed_values: Optional[set[str]] = None



        if allowed_source:

            allowed_source = allowed_source.strip()

            if allowed_source.upper().startswith("VALUE="):

                raw_values = allowed_source[len("VALUE=") :]

                allowed_values = {

                    val.strip().upper()

                    for val in raw_values.split(";")

                    if val and val.strip()

                }

            elif allowed_source.upper().startswith("SHEET="):

                # Parse formats:
                #   SHEET=MySheet
                #   SHEET=MySheet;COLUMN=MyColumn
                #   SHEET=MySheet!MyColumn
                sheet_part = allowed_source[len("SHEET=") :].strip()

                column_name: Optional[str] = None

                # Support bang syntax SHEET=Sheet!Column
                if "!" in sheet_part:
                    sheet_name, after = sheet_part.split("!", 1)
                    sheet_name = sheet_name.strip()
                    column_name = after.strip() or None
                else:
                    # Support key/value pairs SHEET=Sheet;COLUMN=Col
                    parts = [p.strip() for p in sheet_part.split(";") if p and p.strip()]
                    # First token is the sheet name by default
                    sheet_name = parts[0] if parts else ""
                    # Look for COLUMN=... among the remaining parts
                    for token in parts[1:]:
                        if token.upper().startswith("COLUMN=") or token.upper().startswith("COL="):
                            column_name = token.split("=", 1)[1].strip() or None

                if not sheet_name:
                    raise ValueError("Instruction SHEET= invalide: nom de feuille manquant")

                ref_df = reference_cache.get(sheet_name)

                if ref_df is None:

                    ref_df = pd.read_excel(input_file, sheet_name=sheet_name)

                    reference_cache[sheet_name] = ref_df

                # Resolve the column to use
                resolved_column: Optional[str] = None
                if column_name:
                    # Find exact (case-insensitive, trimmed) match
                    for col in ref_df.columns:
                        if normalize(col) == normalize(column_name):
                            resolved_column = col
                            break
                    if resolved_column is None:
                        raise ValueError(
                            f"Colonne '{column_name}' introuvable dans la feuille '{sheet_name}'."
                        )
                else:
                    # Backward-compatible heuristic: try to match the current field name
                    # If not found, fall back to the first non-empty column
                    matching_cols = [col for col in ref_df.columns if normalize(col) == normalize(field)]
                    if len(matching_cols) == 1:
                        resolved_column = matching_cols[0]
                    elif len(matching_cols) > 1:
                        raise ValueError(
                            f"Plusieurs colonnes compatibles avec le champ '{field}' dans la feuille '{sheet_name}'. Precisez COLUMN=."
                        )
                    else:
                        # Fallback: first column with at least one non-empty value
                        non_empty_columns = [
                            col for col in ref_df.columns if any(str(v).strip() for v in ref_df[col].dropna().tolist())
                        ]
                        if non_empty_columns:
                            resolved_column = non_empty_columns[0]
                        else:
                            raise ValueError(
                                f"Aucune colonne valide trouvee dans la feuille '{sheet_name}'."
                            )

                allowed_values = {
                    str(val).strip().upper()
                    for val in ref_df[resolved_column].dropna().tolist()
                    if str(val).strip()
                }



        pattern_value = row.get("Pattern")

        compiled_pattern = re.compile(str(pattern_value)) if isinstance(pattern_value, str) and pattern_value else None



        rules[field] = ValidationRule(

            field=field,

            checked=parse_bool(row.get("Checked")),

            required=parse_bool(row.get("Required")),

            min_length=parse_int(row.get("MinLength")),

            max_length=parse_int(row.get("MaxLength")),

            allowed_values=allowed_values,

            allowed_source=allowed_source,

            pattern=compiled_pattern,

            custom_rule=str(row.get("CustomRule", "")).strip() or None,

        )



    return rules, reference_cache





def fetch_reference_sheet(

    input_file: str,

    cache: dict[str, pd.DataFrame],

    sheet_name: str,

) -> pd.DataFrame:

    sheet = cache.get(sheet_name)

    if sheet is None:

        sheet = pd.read_excel(input_file, sheet_name=sheet_name)

        cache[sheet_name] = sheet

    return sheet





def evaluate_equals_rule(

    rule: ValidationRule,

    row: pd.Series,

    input_file: str,

    reference_cache: Dict[str, pd.DataFrame],

) -> Iterable[str]:

    errors: list[str] = []

    custom_rule = rule.custom_rule or ""

    if not custom_rule.lower().startswith("equals:"):

        return errors



    _, _, payload = custom_rule.partition(":")

    if ";" not in payload:

        return errors



    template_field, ref_column = [part.strip() for part in payload.split(";", 1)]

    join_key_value = str(row.get(rule.field, "")).strip()

    actual_value = str(row.get(template_field, "")).strip().upper()



    allowed_source = rule.allowed_source or ""

    if not allowed_source.upper().startswith("SHEET="):

        return errors



    sheet_name = allowed_source[len("SHEET=") :].strip()

    ref_df = fetch_reference_sheet(input_file, reference_cache, sheet_name)



    join_columns = [

        column for column in ref_df.columns if normalize(column) == normalize(sheet_name)

    ]

    if not join_columns:

        return [

            f"Aucune colonne correspondant a '{sheet_name}' dans la feuille '{sheet_name}'."

        ]



    join_column = join_columns[0]

    matching_rows = ref_df[ref_df[join_column].astype(str).str.strip() == join_key_value]



    if matching_rows.empty or ref_column not in ref_df.columns:

        return [

            f"Valeur '{join_key_value}' introuvable ou colonne '{ref_column}' absente dans '{sheet_name}'."

        ]



    allowed_values: set[str] = set()

    for _, ref_row in matching_rows.iterrows():

        cell_value = ref_row.get(ref_column)

        if pd.isna(cell_value):

            continue

        for candidate in str(cell_value).split(";"):

            candidate = candidate.strip().upper()

            if candidate:

                allowed_values.add(candidate)



    if actual_value and actual_value not in allowed_values:

        formatted_allowed = ", ".join(sorted(allowed_values)) or "(aucune valeur declaree)"

        errors.append(

            f"'{template_field}' doit etre egal a une valeur de '{ref_column}' pour '{rule.field}'='{join_key_value}'."

            f" Valeurs attendues: {formatted_allowed}."

        )

        return errors



    other_values: set[str] = set()

    for _, ref_row in ref_df.iterrows():

        if str(ref_row.get(join_column)).strip() == join_key_value:

            continue

        cell_value = ref_row.get(ref_column)

        if pd.isna(cell_value):

            continue

        for candidate in str(cell_value).split(";"):

            candidate = candidate.strip().upper()

            if candidate:

                other_values.add(candidate)



    if actual_value and actual_value in other_values:

        errors.append(

            f"'{template_field}'='{actual_value}' est deja utilise dans un autre groupe de '{ref_column}'."

        )



    return errors





def evaluate_unique_rule(field: str, row_value: str, counts: dict[str, int]) -> Optional[str]:

    if counts.get(field, 0) > 1:

        return f"'{field}'='{row_value}' n'est pas unique dans la colonne"

    return None





def evaluate_row(

    row: pd.Series,

    rules: dict[str, ValidationRule],

    unique_counts: dict[str, dict[str, int]],

    input_file: str,

    reference_cache: dict[str, pd.DataFrame],

) -> list[str]:

    errors: list[str] = []



    for field, rule in rules.items():

        value = row.get(field)

        value_str = "" if pd.isna(value) else str(value).strip()



        if not rule.checked:

            continue



        if rule.required and not value_str:

            errors.append(f"{field} est requis")

            continue



        if rule.min_length is not None and len(value_str) < rule.min_length:

            errors.append(f"{field} trop court ({len(value_str)} < {rule.min_length})")



        if rule.max_length is not None and len(value_str) > rule.max_length:

            errors.append(f"{field} trop long ({len(value_str)} > {rule.max_length})")



        if rule.allowed_values is not None and value_str.upper() not in rule.allowed_values:

            errors.append(f"Valeur invalide '{value}' pour {field}")



        if rule.pattern and value_str and not rule.pattern.fullmatch(value_str):

            errors.append(f"{field} ne respecte pas le motif {rule.pattern.pattern}")



        if rule.custom_rule:

            custom = rule.custom_rule.strip().lower()

            if custom == "unique":

                unique_error = evaluate_unique_rule(

                    field,

                    value_str,

                    unique_counts.get(field, {}),

                )

                if unique_error:

                    errors.append(unique_error)

            elif custom.startswith("equals:"):

                errors.extend(

                    evaluate_equals_rule(rule, row, input_file, reference_cache)

                )



    return errors





def build_unique_counts(df: pd.DataFrame, rules: dict[str, ValidationRule]) -> dict[str, dict[str, int]]:

    counts: dict[str, dict[str, int]] = {}

    for field, rule in rules.items():

        if rule.custom_rule and rule.custom_rule.strip().lower() == "unique":

            series = df[field].astype(str).str.strip()

            counts[field] = series.value_counts().to_dict()

    return counts





def summarise_errors(df: pd.DataFrame, rules: dict[str, ValidationRule]) -> pd.DataFrame:

    total_rows = len(df)

    summary_records = []

    for field in rules:

        errors_count = df["Errors"].str.contains(field, case=False, na=False).sum()

        error_percentage = f"{round((errors_count / total_rows) * 100, 2)}%" if total_rows else "0%"

        summary_records.append(

            {

                "Field": field,

                "Errors Count": errors_count,

                "Errors %": error_percentage,

            }

        )

    return pd.DataFrame(summary_records)





def add_tables_to_workbook(output_file: str, summary_df: pd.DataFrame, result_df: pd.DataFrame) -> None:

    workbook = load_workbook(output_file)

    summary_sheet = workbook["ErrorSummary"]

    result_sheet = workbook["Result"]



    summary_table = Table(displayName="GlobalStats", ref="A1:B4")

    summary_table.tableStyleInfo = TableStyleInfo(

        name="TableStyleMedium9",

        showRowStripes=True,

    )

    summary_sheet.add_table(summary_table)



    start_row = 6

    end_row = start_row + len(summary_df)

    end_col = get_column_letter(len(summary_df.columns))

    detailed_table = Table(

        displayName="FieldErrors",

        ref=f"A{start_row}:{end_col}{end_row}",

    )

    detailed_table.tableStyleInfo = TableStyleInfo(

        name="TableStyleMedium4",

        showRowStripes=True,

    )

    summary_sheet.add_table(detailed_table)



    result_end_row = len(result_df) + 1

    result_end_col = get_column_letter(len(result_df.columns))

    result_table = Table(

        displayName="ValidationResult",

        ref=f"A1:{result_end_col}{result_end_row}",

    )

    result_table.tableStyleInfo = TableStyleInfo(

        name="TableStyleMedium2",

        showRowStripes=True,

    )

    result_sheet.add_table(result_table)



    workbook.save(output_file)





def write_output(

    input_file: str,

    output_dir: str,

    result_df: pd.DataFrame,

    summary_df: pd.DataFrame,

    valid_flags: list[bool],

) -> str:

    output_filename = Path(input_file).name.replace(".xlsx", " review.xlsx")

    output_path = Path(output_dir) / output_filename



    enriched_df = result_df.copy()

    enriched_df.insert(0, "Valid", [CHECK_MARK if flag else CROSS_MARK for flag in valid_flags])



    total_rows = len(enriched_df)

    valid_rows = sum(valid_flags)

    valid_percentage = f"{round((valid_rows / total_rows) * 100, 2)}%" if total_rows else "0%"



    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        enriched_df.to_excel(writer, sheet_name="Result", index=False)



        metrics = pd.DataFrame(

            {

                "Metric": ["Total Rows", "Valid Rows", "% Valid"],

                "Value": [total_rows, valid_rows, valid_percentage],

            }

        )



        metrics.to_excel(writer, sheet_name="ErrorSummary", startrow=0, index=False)

        summary_df.to_excel(writer, sheet_name="ErrorSummary", startrow=5, index=False)



    add_tables_to_workbook(str(output_path), summary_df, enriched_df)

    return str(output_path)





def generate_result_from_excel(input_file: str, output_dir: str, rules_override: Optional[list[dict[str, object]]] = None) -> str:

    if not os.path.exists(input_file):

        raise FileNotFoundError(f"Fichier introuvable: {input_file}")



    template_df = load_template(input_file)

    override_df: Optional[pd.DataFrame] = None
    if rules_override is not None:
        rows: list[dict[str, object]] = []
        for rule in rules_override:
            field = str(rule.get("field", "")).strip()
            if not field:
                continue

            allowed_type = str(rule.get("allowedType", "instruction")).lower()
            raw_allowed_values = rule.get("allowedValues") or []
            if isinstance(raw_allowed_values, (str, bytes)):
                allowed_values_iter = [str(raw_allowed_values)]
            else:
                allowed_values_iter = [str(value).strip() for value in raw_allowed_values if str(value).strip()]

            joined_values = ";".join(allowed_values_iter)
            if allowed_type == "list" and joined_values:
                allowed_cell = f"VALUE={joined_values}"
            elif allowed_type == "list":
                allowed_cell = ""
            else:
                # Instruction mode: either sheet or custom instruction
                mode = str(rule.get("allowedInstructionMode", "custom")).lower()
                if mode == "sheet":
                    sheet_name = str(rule.get("allowedSheet", "") or "").strip()
                    column_name = str(rule.get("allowedColumn", "") or "").strip()
                    if sheet_name and column_name:
                        allowed_cell = f"SHEET={sheet_name};COLUMN={column_name}"
                    elif sheet_name:
                        allowed_cell = f"SHEET={sheet_name}"
                    else:
                        allowed_cell = ""
                else:
                    allowed_cell = str(rule.get("allowedInstruction", "") or "").strip()

            min_length = rule.get("minLength")
            max_length = rule.get("maxLength")

            rows.append(
                {
                    "Field": field,
                    "Checked": 1 if bool(rule.get("checked")) else 0,
                    "Required": 1 if bool(rule.get("required")) else 0,
                    "MinLength": min_length if min_length is not None else "",
                    "MaxLength": max_length if max_length is not None else "",
                    "AllowedValues": allowed_cell,
                    "Pattern": str(rule.get("pattern", "") or "").strip(),
                    "CustomRule": str(rule.get("customRule", "") or "").strip(),
                }
            )

        if rows:
            override_df = pd.DataFrame(
                rows,
                columns=["Field", "Checked", "Required", "MinLength", "MaxLength", "AllowedValues", "Pattern", "CustomRule"],
            )

    rules, reference_cache = load_validation_rules(input_file, override_df)



    unique_counts = build_unique_counts(template_df, rules)



    valid_flags: list[bool] = []

    error_messages: list[str] = []



    for _, row in template_df.iterrows():

        row_errors = evaluate_row(row, rules, unique_counts, input_file, reference_cache)

        valid_flags.append(len(row_errors) == 0)

        error_messages.append("; ".join(row_errors) if row_errors else "")



    template_df.insert(0, "Errors", error_messages)



    summary_df = summarise_errors(template_df, rules)

    output_path = write_output(input_file, output_dir, template_df, summary_df, valid_flags)

    return output_path





__all__ = ["generate_result_from_excel", "ValidationRule"]




def main() -> None:

    import argparse



    parser = argparse.ArgumentParser(

        description="Valide un template DMF et genere un rapport Excel.",

    )

    parser.add_argument("input", help="Chemin vers le fichier Excel a valider")

    parser.add_argument(

        "--output",

        help="Dossier de sortie (defaut: dossier du fichier d'entree)",

    )



    args = parser.parse_args()

    output_dir = args.output or str(Path(args.input).resolve().parent)

    result = generate_result_from_excel(args.input, output_dir)

    print(f"Validation terminee: {result}")





if __name__ == "__main__":

    main()




